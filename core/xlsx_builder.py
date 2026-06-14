"""
core/xlsx_builder.py
Génère synthese.xlsx depuis synthese.docx.
Extraction Python pure via docx_extractor (sans Mistral).
Détection "Moteur" depuis l'observation.
"""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from core.docx_extractor import extract_eleves_from_docx

# ─── Couleurs ─────────────────────────────────────────────────────────────────
COLORS = {
    "penible": PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid"),  # Orange
    "bilangue": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Bleu ciel
    "basket": PatternFill(start_color="CE93D8", end_color="CE93D8", fill_type="solid"),  # Violet clair
    "moteur": PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid"),  # Vert fluo
    "niveau_F": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),  # Rouge
    "niveau_M": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),  # Orange
    "niveau_B": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),  # Vert clair
    "niveau_TB": PatternFill(start_color="1A5C1A", end_color="1A5C1A", fill_type="solid"),  # Vert foncé
    "aide": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),  # Gris moyen (au lieu de noir)
}

FONT_WHITE = Font(color="FFFFFF", bold=True)
FONT_BLACK = Font(color="000000", bold=True)
FONT_NORMAL = Font()

BORDERS = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ── Colonnes ─────────────────────────────────────────────────────────────────
HEADERS = [
    "Nom",
    "Prénom",
    "Sexe",
    "Niveau",
    "Ville",
    "Bilangue",
    "Basket",
    "Moteur",
    "Pénible",
    "Aide",
    "À séparer de",
    "À regrouper avec",
]

def _clean_nom(nom: str) -> str:
    """Nettoie le nom des emojis (basket, etc.)."""
    emojis_to_remove = ["🏀", "⛹️", "⛹", "🏀"]
    cleaned = nom
    for emoji in emojis_to_remove:
        cleaned = cleaned.replace(emoji, "").strip()
    return cleaned

def _normalize_ville(ville: str) -> str:
    """Normalise le nom de la ville.
    - Sainte-Marie-aux-Chênes → SMAC
    - Montois-la-Montagne → Montois
    - Saint-Privat-la-Montagne → Saint-Privat
    """
    if not ville:
        return "SMAC"
    
    ville_lower = ville.lower().strip()
    
    # Supprimer les espaces multiples
    ville_lower = " ".join(ville_lower.split())
    
    # Cas spéciaux pour les villes composées
    if "montois-la-montagne" in ville_lower or "montois la montagne" in ville_lower:
        return "Montois"
    
    if "saint-privat-la-montagne" in ville_lower or "saint privat la montagne" in ville_lower:
        return "Saint-Privat"
    
    # SMAC (Sainte-Marie-aux-Chênes)
    if any(s in ville_lower for s in [
        "sainte-marie-aux-chênes", 
        "ste marie aux chenes", 
        "sainte marie aux chenes", 
        "smac",
        "ste marie-aux-chenes"
    ]):
        return "SMAC"
    
    # Retourner la ville telle quelle si pas de correspondance
    return ville

def _get_max_text_length(value) -> int:
    """Calcule la longueur maximale du texte pour ajuster la colonne."""
    if value is None:
        return 0
    text = str(value)
    return len(text)

def build_synthese_xlsx(docx_path: str | Path, output_path: str | Path) -> None:
    """
    Génère synthese.xlsx depuis synthese.docx.
    """
    eleves = extract_eleves_from_docx(docx_path)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    
    # ── TITRE PRINCIPAL ───────────────────────────────────────────────────────
    TITRE = "📊 Synthèse des effectifs 6e pour la prochaine rentrée"
    
    # Fusionner les cellules pour le titre
    ws.merge_cells('A1:K1')
    titre_cell = ws['A1']
    titre_cell.value = TITRE
    titre_cell.font = Font(bold=True, size=14, color="1A5C1A")  # Vert foncé
    titre_cell.alignment = Alignment(horizontal="center", vertical="center")
    titre_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")  # Vert très clair
    
    # Ajuster la hauteur de la ligne du titre
    ws.row_dimensions[1].height = 35
    
    # ── En-têtes (ligne 2) ────────────────────────────────────────────────────
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.border = BORDERS
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Hauteur des en-têtes
    ws.row_dimensions[2].height = 30
    
    # ── Données (à partir de ligne 3) ─────────────────────────────────────────
    for row_idx, eleve in enumerate(eleves, 3):
        # Nettoyage du nom (suppression emojis)
        nom_clean = _clean_nom(eleve.nom)
        
        # Normalisation de la ville
        ville_clean = _normalize_ville(eleve.commune)
        
        # Données de base
        data = {
            "Nom": nom_clean,
            "Prénom": eleve.prenom,
            "Sexe": "G" if eleve.sexe.value == "G" else "F",
            "Niveau": eleve.niveau.value,  # <--- AJOUTÉ ICI
            "Ville": ville_clean.upper(),
            "Bilangue": "OUI" if eleve.est_bilangue else "",
            "Basket": "OUI" if eleve.est_basket else "",
            "Moteur": "OUI" if "moteur" in (eleve.observation or "").lower() else "",
            "Pénible": "OUI" if eleve.est_penible else "",
            "Aide": eleve.aide.value if eleve.aide.value else "",
            "À séparer de": ", ".join(eleve.a_separer_de),
            "À regrouper avec": ", ".join(eleve.a_regrouper_avec),
        }
        
        # Écriture des cellules
        for col_idx, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data[header])
            cell.border = BORDERS
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            # Alignement : Nom et Prénom à gauche, le reste au centre
            if header in ("Nom", "Prénom"):
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
            
                        # Appliquer les couleurs conditionnelles
            valeur = data[header]  # ← CORRECTION : utiliser data[header] au lieu de val
            
            if header == "Niveau":
                if valeur == "TB":
                    cell.fill = PatternFill(start_color="00CC00", end_color="00CC00", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                elif valeur == "B":
                    cell.fill = PatternFill(start_color="33FF33", end_color="33FF33", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)
                elif valeur == "M":
                    cell.fill = PatternFill(start_color="FF9933", end_color="FF9933", fill_type="solid")
                    cell.font = Font(color="000000", bold=True)
                elif valeur == "F":
                    cell.fill = PatternFill(start_color="FF3333", end_color="FF3333", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
            
            elif header == "Pénible" and valeur == "OUI":
                cell.fill = COLORS["penible"]
                cell.font = FONT_WHITE
            
            elif header == "Bilangue" and valeur == "OUI":
                cell.fill = COLORS["bilangue"]
                cell.font = FONT_BLACK
            
            elif header == "Basket" and valeur == "OUI":
                cell.fill = COLORS["basket"]
                cell.font = FONT_BLACK
            
            elif header == "Moteur" and valeur == "OUI":
                cell.fill = COLORS["moteur"]
                cell.font = FONT_BLACK
            
            elif header == "Aide" and valeur:
                cell.fill = COLORS["aide"]
                cell.font = FONT_WHITE
            
            # Police par défaut
            if not cell.font.color.rgb:
                cell.font = Font(size=10)
    
    # ── Largeurs colonnes (ajustement automatique avec minimum) ───────────────
    column_widths = {
        "A": 25,  # Nom
        "B": 18,  # Prénom
        "C": 10,  # Sexe (élargi pour la flèche de tri)
        "D": 18,  # Ville
        "E": 14,  # Bilangue (élargi pour la flèche de tri)
        "F": 12,  # Basket (élargi pour la flèche de tri)
        "G": 12,  # Moteur (élargi pour la flèche de tri)
        "H": 12,  # Pénible (élargi pour la flèche de tri)
        "I": 14,  # Aide (élargi pour la flèche de tri)
        "J": 30,  # À séparer de
        "K": 30,  # À regrouper avec
    }
    
    # Ajustement dynamique basé sur le contenu
    for col_idx, header in enumerate(HEADERS, 1):
        col_letter = get_column_letter(col_idx)
        max_length = len(header) + 3  # +3 pour le padding (espace pour la flèche de tri)
        
        # Parcourir toutes les lignes de données
        for row in range(3, len(eleves) + 3):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length
        
        # Appliquer la largeur (avec un minimum défini)
        min_width = column_widths.get(col_letter, 15)
        new_width = max(min_width, min(max_length, 40))  # Maximum 40 pour éviter trop large
        ws.column_dimensions[col_letter].width = new_width
    
    # ── Hauteurs de lignes ───────────────────────────────────────────────────
    for row in range(3, len(eleves) + 3):
        ws.row_dimensions[row].height = 25  # Hauteur confortable
    
    # ── Filtres automatiques (en-têtes triables) ─────────────────────────────
    ws.auto_filter.ref = f"A2:K{len(eleves) + 2}"
    
    # ── Sauvegarde ────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"✅ synthese.xlsx généré : {output_path}")


# ─── API publique ──────────────────────────────────────────────────────────────

def generate_synthese_xlsx(docx_path: str | Path, output_dir: str | Path = ".") -> Path:
    """
    Génère synthese.xlsx depuis synthese.docx.
    Retourne le chemin du fichier généré.
    """
    docx_path = Path(docx_path)
    output_dir = Path(output_dir)
    output_path = output_dir / "synthese.xlsx"
    
    build_synthese_xlsx(docx_path, output_path)
    
    return output_path