"""
core/distributor.py
Algorithme de répartition des élèves en classes de 6e.
Force l'équilibre des effectifs + respecte les contraintes.
"""
from __future__ import annotations
import random
from pathlib import Path
from dataclasses import dataclass, field
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Couleurs ─────────────────────────────────────────────────────────────────
COLORS = {
    "penible": PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid"),
    "bilangue": PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
    "basket": PatternFill(start_color="CE93D8", end_color="CE93D8", fill_type="solid"),
    "moteur": PatternFill(start_color="39FF14", end_color="39FF14", fill_type="solid"),
    "niveau_F": PatternFill(start_color="F44336", end_color="F44336", fill_type="solid"),
    "niveau_M": PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid"),
    "niveau_B": PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"),
    "niveau_TB": PatternFill(start_color="1A5C1A", end_color="1A5C1A", fill_type="solid"),
    "aide": PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid"),
}
FONT_WHITE = Font(color="FFFFFF", bold=True, size=10)
FONT_BLACK = Font(color="000000", bold=True, size=10)

# Bordures fines pour les cellules de données
BORDERS = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# Bordures épaisses pour le résumé
BORDERS_THICK = Border(
    left=Side(style="medium", color="333333"),
    right=Side(style="medium", color="333333"),
    top=Side(style="medium", color="333333"),
    bottom=Side(style="medium", color="333333"),
)

TAB_COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B59B6", "E74C3C", "1ABC9C",
]

HEADERS = [
    "Nom", "Prénom", "Sexe",
    "Niveau",
    "Ville",
    "Bilangue", "Basket", "Moteur", "Pénible",
    "Aide", "À séparer de", "À regrouper avec",
]

# ── Modèles de données ───────────────────────────────────────────────────────
@dataclass
class Eleve:
    nom: str
    prenom: str
    sexe: str
    ville: str
    est_bilangue: bool
    est_basket: bool
    est_moteur: bool
    est_penible: bool
    aide: str
    a_separer_de: list[str]
    a_regrouper_avec: list[str]
    niveau: str = "?"
    classe_attribuee: str = ""

    @property
    def identifiant(self) -> str:
        return f"{self.nom.upper()} {self.prenom.upper()}"

@dataclass
class Classe:
    nom: str
    est_bilangue: bool = False
    est_basket: bool = False
    eleves: list[Eleve] = field(default_factory=list)

    @property
    def nb_garcons(self) -> int:
        return sum(1 for e in self.eleves if e.sexe == "G")

    @property
    def nb_filles(self) -> int:
        return sum(1 for e in self.eleves if e.sexe == "F")

    @property
    def effectif(self) -> int:
        return len(self.eleves)

    def compteurs_niveaux(self) -> dict:
        c = {"TB": 0, "B": 0, "M": 0, "F": 0}
        for e in self.eleves:
            if e.niveau in c:
                c[e.niveau] += 1
        return c

    def contient_ennemi(self, eleve: Eleve) -> bool:
        ids_classe = {e.identifiant for e in self.eleves}
        for ennemi in eleve.a_separer_de:
            ennemi_upper = ennemi.upper().strip()
            if any(ennemi_upper in id_classe for id_classe in ids_classe):
                return True
        return False

    def contient_ami(self, eleve: Eleve) -> bool:
        ids_classe = {e.identifiant for e in self.eleves}
        for ami in eleve.a_regrouper_avec:
            ami_upper = ami.upper().strip()
            if any(ami_upper in id_classe for id_classe in ids_classe):
                return True
        return False

# ─── Lecture du XLSX source ───────────────────────────────────────────────────
def lire_eleves_depuis_xlsx(xlsx_path: str | Path) -> list[Eleve]:
    """Lit le synthese.xlsx et retourne la liste des élèves."""
    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    
    eleves = []
    
    # min_row=3 car ligne 1=titre, ligne 2=en-têtes
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not row[0]:
            continue
        
        nom = str(row[0] or "").strip()
        prenom = str(row[1] or "").strip()
        
        if not nom or nom.lower() == "nom":
            continue
        
        eleves.append(Eleve(
            nom=nom,
            prenom=prenom,
            sexe=str(row[2] or "F").strip(),
            niveau=str(row[3] or "?").strip().upper(),
            ville=str(row[4] or "").strip(),
            est_bilangue=str(row[5] or "").strip().upper() == "OUI",
            est_basket=str(row[6] or "").strip().upper() == "OUI",
            est_moteur=str(row[7] or "").strip().upper() == "OUI",
            est_penible=str(row[8] or "").strip().upper() == "OUI",
            aide=str(row[9] or "").strip(),
            a_separer_de=[n.strip() for n in str(row[10] or "").split(",") if n.strip()],
            a_regrouper_avec=[n.strip() for n in str(row[11] or "").split(",") if n.strip()],
        ))
    
    wb.close()
    return eleves

# ─── Moteur de répartition ────────────────────────────────────────────────────
def repartir_eleves(
    eleves: list[Eleve],
    classes: list[Classe],
    effectif_cible: int = 28,
) -> list[Classe]:
    """
    Algorithme de répartition par phases avec ÉQUILIBRE FORCÉ.
    """
    random.shuffle(eleves)
    
    nb_classes = len(classes)
    effectif_ideal = len(eleves) / nb_classes
    effectif_max = int(effectif_ideal) + 2
    
    print(f"[DEBUG] {len(eleves)} élèves / {nb_classes} classes = {effectif_ideal:.1f} idéal/classe")
    print(f"[DEBUG] Effectif max autorisé : {effectif_max}")

    # ── PHASE 1 : Basket et Bilangue (avec round-robin pour équilibrer) ─────
    basket_eleves = [e for e in eleves if e.est_basket]
    bilangue_eleves = [e for e in eleves if e.est_bilangue and not e.est_basket]
    autres_eleves = [e for e in eleves if not e.est_basket and not e.est_bilangue]

    # 1. Basket → Classe Basket (round-robin si plusieurs classes basket)
    classes_basket = [c for c in classes if c.est_basket]
    if classes_basket:
        for i, e in enumerate(basket_eleves):
            c = classes_basket[i % len(classes_basket)]
            c.eleves.append(e)
            e.classe_attribuee = c.nom
    else:
        # Pas de classe basket : on les disperse
        autres_eleves.extend(basket_eleves)

    # 2. Bilangue → Classes bilangues (round-robin)
    classes_bilangues = [c for c in classes if c.est_bilangue]
    if classes_bilangues:
        for i, e in enumerate(bilangue_eleves):
            c = classes_bilangues[i % len(classes_bilangues)]
            c.eleves.append(e)
            e.classe_attribuee = c.nom
    else:
        autres_eleves.extend(bilangue_eleves)

    # ── PHASE 2 : Pénibles et Moteurs (séparation FORCÉE dans TOUTES les classes) ──
    penibles = [e for e in autres_eleves if e.est_penible]
    moteurs = [e for e in autres_eleves if e.est_moteur and not e.est_penible]
    reste = [e for e in autres_eleves if not e.est_penible and not e.est_moteur]

    # Calculer combien de pénibles par classe (répartition équitable)
    nb_penibles = len(penibles)
    penibles_par_classe = max(1, nb_penibles // nb_classes)  # Au moins 1 par classe
    
    # Trier les pénibles par nombre d'ennemis (les plus "difficiles" d'abord)
    penibles.sort(key=lambda e: len(e.a_separer_de), reverse=True)
    
    # Répartir les pénibles de manière équitable
    for e in penibles:
        # Trouver la classe qui a le MOINS de pénibles et qui n'a pas d'ennemi
        meilleure = None
        min_penibles = float('inf')
        
        for c in sorted(classes, key=lambda x: x.effectif):
            nb_penibles_classe = sum(1 for el in c.eleves if el.est_penible)
            
            # Priorité aux classes avec le moins de pénibles
            if nb_penibles_classe < min_penibles and not c.contient_ennemi(e):
                if c.effectif < effectif_max:
                    meilleure = c
                    min_penibles = nb_penibles_classe
        
        if meilleure:
            meilleure.eleves.append(e)
            e.classe_attribuee = meilleure.nom
        else:
            # Fallback : classe la moins remplie
            c = min(classes, key=lambda x: x.effectif)
            c.eleves.append(e)
            e.classe_attribuee = c.nom

    # Même logique pour les moteurs
    moteurs.sort(key=lambda e: len(e.a_separer_de), reverse=True)
    
    for e in moteurs:
        meilleure = None
        min_moteurs = float('inf')
        
        for c in sorted(classes, key=lambda x: x.effectif):
            nb_moteurs_classe = sum(1 for el in c.eleves if el.est_moteur)
            
            if nb_moteurs_classe < min_moteurs and not c.contient_ennemi(e):
                if c.effectif < effectif_max:
                    meilleure = c
                    min_moteurs = nb_moteurs_classe
        
        if meilleure:
            meilleure.eleves.append(e)
            e.classe_attribuee = meilleure.nom
        else:
            c = min(classes, key=lambda x: x.effectif)
            c.eleves.append(e)
            e.classe_attribuee = c.nom

    # ── PHASE 3 : Équilibre (Genre, Niveaux, Regroupement) ─────────────────
    # Tri par niveau (du plus rare au plus commun)
    niveau_order = {"F": 0, "M": 1, "B": 2, "TB": 3}
    reste.sort(key=lambda e: niveau_order.get(e.niveau, 99))

    for e in reste:
        scores = []
        for c in classes:
            score = 0.0

            # 1. PÉNALITÉ EXPONENTIELLE si dépassement effectif max
            if c.effectif >= effectif_max:
                score -= 10000  # Interdiction quasi-totale
            elif c.effectif >= effectif_ideal:
                score -= 500 * (c.effectif - effectif_ideal)  # Pénalité progressive

            # 2. BONUS si classe sous-remplie (pour équilibrer)
            if c.effectif < effectif_ideal:
                score += 200 * (effectif_ideal - c.effectif)

            # 3. Équilibre Garçons/Filles (tendre vers 50/50)
            if c.effectif > 0:
                ratio_g = c.nb_garcons / c.effectif
                if e.sexe == "G" and ratio_g < 0.5:
                    score += 50
                elif e.sexe == "F" and ratio_g > 0.5:
                    score += 50

            # 4. Équilibre des niveaux (TRÈS IMPORTANT)
            compteurs = c.compteurs_niveaux()
            min_niveau = min(compteurs.values()) if compteurs else 0
            if compteurs.get(e.niveau, 0) <= min_niveau:
                score += 100  # Augmenté de 30 à 100 pour forcer l'équilibre

            # 5. Bonus "À regrouper avec"
            if c.contient_ami(e):
                score += 100

            # 6. Pénalité "À séparer de" (hard constraint)
            if c.contient_ennemi(e):
                score -= 5000

            scores.append((score, c))

        scores.sort(key=lambda x: x[0], reverse=True)
        meilleure_classe = scores[0][1]
        meilleure_classe.eleves.append(e)
        e.classe_attribuee = meilleure_classe.nom

    # Vérification finale
    for c in classes:
        compteurs = c.compteurs_niveaux()
        nb_pen = sum(1 for e in c.eleves if e.est_penible)
        print(f"[DEBUG] {c.nom} : {c.effectif} élèves ({c.nb_garcons}G/{c.nb_filles}F) - "
              f"TB/B/M/F: {compteurs['TB']}/{compteurs['B']}/{compteurs['M']}/{compteurs['F']} - "
              f"Pénibles: {nb_pen}")
    
    total_reparti = sum(c.effectif for c in classes)
    print(f"[DEBUG] Total réparti : {total_reparti} / {len(eleves)}")

    return classes

# ─── Génération Excel ─────────────────────────────────────────────────────────
def _appliquer_couleurs(cell, col_idx: int, val):
    """Applique les couleurs et bordures à une cellule."""
    cell.border = BORDERS
    cell.alignment = Alignment(vertical="center", horizontal="left" if col_idx <= 2 else "center")

    if col_idx == 4:  # Niveau
        if val == "TB":
            cell.fill = PatternFill(start_color="1A5C1A", end_color="1A5C1A", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
        elif val == "B":
            cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            cell.font = Font(color="000000", bold=True, size=10)
        elif val == "M":
            cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
            cell.font = Font(color="000000", bold=True, size=10)
        elif val == "F":
            cell.fill = PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True, size=10)
            
    elif col_idx == 9 and val == "OUI":  # Pénible
        cell.fill = COLORS["penible"]
        cell.font = FONT_WHITE
    elif col_idx == 6 and val == "OUI":  # Bilangue
        cell.fill = COLORS["bilangue"]
        cell.font = FONT_BLACK
    elif col_idx == 7 and val == "OUI":  # Basket
        cell.fill = COLORS["basket"]
        cell.font = FONT_BLACK
    elif col_idx == 8 and val == "OUI":  # Moteur
        cell.fill = COLORS["moteur"]
        cell.font = Font(color="000000", bold=True, size=10)
    elif col_idx == 10 and val:  # Aide
        cell.fill = COLORS["aide"]
        cell.font = FONT_WHITE

def generer_xlsx_repartition(
    classes: list[Classe],
    output_path: str | Path,
    annee: str = "2026",
) -> None:
    """
    Génère le fichier repartition_[annee].xlsx avec onglets par classe.
    """
    wb = Workbook()

    # ── Onglet "Synthèse" ───────────────────────────────────────────────────
    ws_global = wb.active
    ws_global.title = "Synthèse"
    ws_global.sheet_properties.tabColor = "4472C4"
    ws_global.sheet_view.showGridLines = False

    # TITRE
    TITRE = "📊 Répartition des effectifs 6e pour la prochaine rentrée"
    ws_global.merge_cells('A1:M1')
    titre_cell = ws_global['A1']
    titre_cell.value = TITRE
    titre_cell.font = Font(bold=True, size=14, color="1A5C1A")
    titre_cell.alignment = Alignment(horizontal="center", vertical="center")
    titre_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    titre_cell.border = Border(
        left=Side(style="thick", color="1A5C1A"),
        right=Side(style="thick", color="1A5C1A"),
        top=Side(style="thick", color="1A5C1A"),
        bottom=Side(style="thick", color="1A5C1A"),
    )
    ws_global.row_dimensions[1].height = 35

    headers_complets = HEADERS + ["Classe de 6e"]

    for col, h in enumerate(headers_complets, 1):
        cell = ws_global.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        cell.border = BORDERS
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws_global.row_dimensions[2].height = 30

    tous_eleves = []
    for c in classes:
        tous_eleves.extend(sorted(c.eleves, key=lambda x: x.nom))

    row_idx = 3
    for e in tous_eleves:
        data = [
            e.nom, e.prenom, e.sexe,
            e.niveau,
            e.ville,
            "OUI" if e.est_bilangue else "",
            "OUI" if e.est_basket else "",
            "OUI" if e.est_moteur else "",
            "OUI" if e.est_penible else "",
            e.aide,
            ", ".join(e.a_separer_de),
            ", ".join(e.a_regrouper_avec),
            e.classe_attribuee,
        ]
        for col_idx, val in enumerate(data, 1):
            cell = ws_global.cell(row=row_idx, column=col_idx, value=val)
            _appliquer_couleurs(cell, col_idx, val)
        row_idx += 1

    ws_global.auto_filter.ref = f"A2:M{row_idx - 1}"

    largeurs = [20, 15, 6, 10, 18, 12, 12, 12, 12, 14, 25, 25, 15]
    for i, w in enumerate(largeurs, 1):
        ws_global.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════════════════
    # ONGLETS PAR CLASSE (UNE SEULE BOUCLE !)
    # ═══════════════════════════════════════════════════════════════════════════
    for idx_classe, c in enumerate(classes):
        ws = wb.create_sheet(title=c.nom)
        couleur_onglet = TAB_COLORS[idx_classe % len(TAB_COLORS)]
        ws.sheet_properties.tabColor = couleur_onglet
        ws.sheet_view.showGridLines = False
        
        # ── LIGNE 1 : TITRE DE LA CLASSE ─────────────────────────────────
        ws.merge_cells('A1:L1')
        titre_cell = ws['A1']
        titre_cell.value = f"Classe de {c.nom}"
        titre_cell.font = Font(bold=True, size=16, color=couleur_onglet, name="Calibri")
        titre_cell.alignment = Alignment(horizontal="center", vertical="center")
        titre_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        ws.row_dimensions[1].height = 35
        
        # Bordure épaisse de la couleur de l'onglet
        border_titre = Border(
            left=Side(style="thick", color=couleur_onglet),
            right=Side(style="thick", color=couleur_onglet),
            top=Side(style="thick", color=couleur_onglet),
            bottom=Side(style="thick", color=couleur_onglet),
        )
        
        # Appliquer la bordure à TOUTES les cellules de la fusion pour qu'elle s'affiche sur toute la largeur
        for col in range(1, 13):
            ws.cell(row=1, column=col).border = border_titre
        
        # ── LIGNE 2 : LIGNE VIERGE DE SÉPARATION ──────────────────────────
        ws.row_dimensions[2].height = 8  # Hauteur réduite
        
        # ── LIGNE 3 : EN-TÊTES DU TABLEAU ─────────────────────────────────
        for col, h in enumerate(HEADERS, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            cell.border = BORDERS
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.row_dimensions[3].height = 30

        # ── LIGNES 4+ : DONNÉES DES ÉLÈVES ────────────────────────────────
        for r_idx, e in enumerate(sorted(c.eleves, key=lambda x: x.nom), 4):
            data = [
                e.nom, e.prenom, e.sexe,
                e.niveau,
                e.ville,
                "OUI" if e.est_bilangue else "",
                "OUI" if e.est_basket else "",
                "OUI" if e.est_moteur else "",
                "OUI" if e.est_penible else "",
                e.aide,
                ", ".join(e.a_separer_de),
                ", ".join(e.a_regrouper_avec),
            ]
            for col_idx, val in enumerate(data, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                _appliquer_couleurs(cell, col_idx, val)

        # Filtres automatiques
        ws.auto_filter.ref = f"A3:L{len(c.eleves) + 3}"

        # Largeurs colonnes
        for i, w in enumerate(largeurs[:12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ═══════════════════════════════════════════════════════════════════
        # RÉSUMÉ STATISTIQUE STYLÉ
        # ═══════════════════════════════════════════════════════════════════
        row_resume = len(c.eleves) + 6  # Laisse 2 lignes vides après le tableau
        
        # Ligne vide de séparation
        ws.row_dimensions[row_resume - 1].height = 15
        
        # Titre "STATISTIQUES" fusionné
        ws.merge_cells(f'A{row_resume}:L{row_resume}')
        titre_stats = ws.cell(row=row_resume, column=1, value="📊 STATISTIQUES DE LA CLASSE")
        titre_stats.font = Font(bold=True, size=14, color="FFFFFF", name="Calibri")
        titre_stats.alignment = Alignment(horizontal="center", vertical="center")
        titre_stats.fill = PatternFill(start_color=couleur_onglet, end_color=couleur_onglet, fill_type="solid")
        border_stats = Border(
            left=Side(style="thick", color=couleur_onglet),
            right=Side(style="thick", color=couleur_onglet),
            top=Side(style="thick", color=couleur_onglet),
            bottom=Side(style="thick", color=couleur_onglet),
        )
        for col in range(1, 13):
            ws.cell(row=row_resume, column=col).border = border_stats
        ws.row_dimensions[row_resume].height = 30
        
        # Ligne vide
        row_resume += 1
        
        # Tableau de statistiques (2 colonnes : Label | Valeur)
        stats_data = [
            ("Effectif total", c.effectif, "1A5C1A"),  # Vert foncé
            ("Garçons", c.nb_garcons, "4472C4"),       # Bleu
            ("Filles", c.nb_filles, "E91E63"),         # Rose
        ]
        
        compteurs = c.compteurs_niveaux()
        stats_data.append(("TB / B / M / F", f"{compteurs['TB']} / {compteurs['B']} / {compteurs['M']} / {compteurs['F']}", "FF9800"))  # Orange
        
        for label, valeur, couleur in stats_data:
            # Cellule label
            cell_label = ws.cell(row=row_resume, column=1, value=label)
            cell_label.font = Font(bold=True, size=12, color="333333", name="Calibri")
            cell_label.alignment = Alignment(horizontal="left", vertical="center")
            cell_label.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            cell_label.border = BORDERS_THICK
            
            # Fusionner les colonnes A à E pour le label
            ws.merge_cells(f'A{row_resume}:E{row_resume}')
            
            # Cellule valeur
            cell_valeur = ws.cell(row=row_resume, column=6, value=valeur)
            cell_valeur.font = Font(bold=True, size=14, color=couleur, name="Calibri")
            cell_valeur.alignment = Alignment(horizontal="center", vertical="center")
            cell_valeur.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell_valeur.border = BORDERS_THICK
            
            # Fusionner les colonnes F à L pour la valeur
            ws.merge_cells(f'F{row_resume}:L{row_resume}')
            
            ws.row_dimensions[row_resume].height = 28
            row_resume += 1
        
        # Ligne vide finale
        ws.row_dimensions[row_resume].height = 10

    # ── SAUVEGARDE ─────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"✅ repartition.xlsx généré : {output_path}")